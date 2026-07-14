"""Create a publication-safe copy containing one anonymized proposal sheet."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


SAFE_STUDENT = re.compile(r"Student\s+\d+", re.IGNORECASE)
SAFE_EXTERNAL = re.compile(r"[A-Z0-9]+(?:-[A-Z0-9]+)*")


def find_header(sheet, label: str, max_rows: int = 10) -> tuple[int, int]:
    for row in range(1, min(max_rows, sheet.max_row) + 1):
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, str) and value.strip().upper() == label.upper():
                return row, column
    raise ValueError(f"Header '{label}' was not found in worksheet '{sheet.title}'.")


def sanitize(source: Path, output: Path, worksheet: str) -> None:
    workbook = load_workbook(source, data_only=False, keep_links=False)
    if worksheet not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{worksheet}' was not found in {source.name}.")

    for sheet_name in list(workbook.sheetnames):
        if sheet_name != worksheet:
            del workbook[sheet_name]
    sheet = workbook[worksheet]

    student_header_row, student_column = find_header(sheet, "Nama Pelajar")
    external_header_row, external_column = find_header(sheet, "LUAR")
    for row in range(student_header_row + 1, sheet.max_row + 1):
        student = sheet.cell(row, student_column).value
        if isinstance(student, str) and student.strip() and not SAFE_STUDENT.fullmatch(student.strip()):
            raise ValueError(f"Non-anonymized student value at {sheet.title}!{sheet.cell(row, student_column).coordinate}")

    for row in range(external_header_row + 1, sheet.max_row + 1):
        cell = sheet.cell(row, external_column)
        value = cell.value
        if isinstance(value, str) and value.strip():
            cleaned = value.strip()
            if cleaned.casefold() == "rozniza":
                cell.value = "R-PPIP"
            elif not SAFE_EXTERNAL.fullmatch(cleaned):
                raise ValueError(f"Non-anonymized external-panel value at {sheet.title}!{cell.coordinate}")

    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.comment = None
            if cell.hyperlink:
                cell._hyperlink = None

    properties = workbook.properties
    properties.creator = "PTPM USM"
    properties.lastModifiedBy = "PTPM USM"
    properties.title = "PTPM Proposal Panel Data"
    properties.subject = "Anonymized proposal panel appointments"
    properties.keywords = "PTPM, proposal panel, anonymized"
    properties.description = "Public-safe anonymized dashboard source data"
    workbook._external_links = []
    workbook.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("worksheet")
    args = parser.parse_args()
    sanitize(args.source, args.output, args.worksheet)


if __name__ == "__main__":
    main()
